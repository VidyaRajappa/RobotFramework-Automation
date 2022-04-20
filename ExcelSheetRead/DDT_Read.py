import openpyxl

wk = openpyxl.load_workbook("C:/Users/VidyashreeR/PycharmProjects/V1/Excel Program data/ReadData100.xlsx")

#create object of sheet
sh = wk['Sheet1']

#Finding maximum rows in the sheet
print(sh.max_row)

#fetch Data from any cell
cell = sh.cell(1,1)
print(cell.value)