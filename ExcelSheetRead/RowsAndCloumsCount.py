import openpyxl

wk = openpyxl.load_workbook("C:/Users/VidyashreeR/Desktop/AWT/AWT Test Cases/ModellingAppFlow.xlsx")

sh = wk['Sheet1']

#Fetch value of rows and columns
rows = sh.max_row
columns = sh.max_column

print("Total Rows are - " + str(rows))
print("Total Columns are - " + str(columns))

