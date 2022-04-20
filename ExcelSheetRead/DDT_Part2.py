import openpyxl

wk = openpyxl.load_workbook("C:/Users/VidyashreeR/PycharmProjects/V1/Excel Program data/ReadData100.xlsx")

#create method to fetch number of rows

def Fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

#create one more method to fetch data from cell
def data_from_cell(Sheetname,row,cell):
    sh = wk[Sheetname]
    cell = sh.cell(row,cell)
    return cell.value

#call the methods
print(Fetch_number_of_rows("Sheet1"))
print(data_from_cell("Sheet1",1,1))